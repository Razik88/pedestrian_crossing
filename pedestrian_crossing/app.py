import os
import cv2
import numpy as np
import sqlite3
from datetime import datetime
from flask import Flask, request, render_template, jsonify, send_from_directory, send_file
from ultralytics import YOLO
import io
import traceback

# --- PDF ---
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader

# --- Excel ---
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from io import BytesIO

app = Flask(__name__)

# ---------- БАЗА ДАННЫХ ----------
def init_db():
    conn = sqlite3.connect('history.db')
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS detections (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    timestamp TEXT,
                    original_file TEXT,
                    processed_file TEXT,
                    pedestrian_count INTEGER,
                    total_detected INTEGER
                )''')
    conn.commit()
    conn.close()

init_db()

# ---------- МОДЕЛИ ----------
# Модель для людей (стандартная детекция)
model_people = YOLO('yolov8n.pt')

# Модель для зебры (сегментация)
model_zebra_path = 'zebra.pt'
if os.path.exists(model_zebra_path):
    model_zebra = YOLO(model_zebra_path)
    print("✅ Модель зебры загружена")
else:
    model_zebra = None
    print("⚠️ zebra.pt не найдена, зебра не будет определяться")

STATIC_FOLDER = 'static'
os.makedirs(STATIC_FOLDER, exist_ok=True)

# ---------- ГЛАВНАЯ ----------
@app.route('/')
def index():
    return render_template('index.html')

# ---------- ОБРАБОТКА ----------
@app.route('/process', methods=['POST'])
def process_image():
    try:
        if 'image' not in request.files:
            return jsonify({'error': 'Файл не загружен'}), 400
        file = request.files['image']
        if file.filename == '':
            return jsonify({'error': 'Пустое имя файла'}), 400

        file_bytes = np.frombuffer(file.read(), np.uint8)
        img = cv2.imdecode(file_bytes, cv2.IMREAD_COLOR)
        if img is None:
            return jsonify({'error': 'Не удалось прочитать изображение'}), 400

        height, width = img.shape[:2]

        # ---------- 1. ДЕТЕКЦИЯ ЛЮДЕЙ ----------
        results_people = model_people(img)
        pedestrians = []
        if results_people[0].boxes is not None:
            for box in results_people[0].boxes:
                if int(box.cls[0].item()) == 0:  # person
                    x1, y1, x2, y2 = map(int, box.xyxy[0].tolist())
                    cx = (x1 + x2) // 2
                    cy = (y1 + y2) // 2
                    pedestrians.append({
                        'bbox': (x1, y1, x2, y2),
                        'center': (cx, cy),
                        'conf': box.conf[0].item()
                    })

        total = len(pedestrians)

        # ---------- 2. СЕГМЕНТАЦИЯ ЗЕБРЫ ----------
        zebra_mask = None
        zebra_bbox = None
        if model_zebra is not None:
            results_zebra = model_zebra(img)
            if results_zebra[0].masks is not None and len(results_zebra[0].masks.data) > 0:
                mask = results_zebra[0].masks.data[0].cpu().numpy()
                mask = (mask > 0.5).astype(np.uint8) * 255
                mask = cv2.resize(mask, (width, height))
                zebra_mask = mask

                contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
                if contours:
                    x, y, w, h = cv2.boundingRect(max(contours, key=cv2.contourArea))
                    zebra_bbox = (x, y, w, h)

        # ---------- 3. ПОДСЧЁТ ЛЮДЕЙ НА ЗЕБРЕ ----------
        on_crosswalk = []
        if zebra_mask is not None:
            for p in pedestrians:
                x1, y1, x2, y2 = p['bbox']
                cx = (x1 + x2) // 2
                cy = (y1 + y2) // 2
                points = [
                    (cx, cy),
                    (x1, y2),
                    (x2, y2),
                    ((x1 + x2) // 2, y2)
                ]
                on_zebra = False
                for px, py in points:
                    if 0 <= px < width and 0 <= py < height and zebra_mask[py, px] == 255:
                        on_zebra = True
                        break
                if on_zebra:
                    on_crosswalk.append(p)
        else:
            # Запасной вариант (если зебра не найдена)
            line_y = height // 2
            for p in pedestrians:
                if abs(p['center'][1] - line_y) <= 20:
                    on_crosswalk.append(p)

        count = len(on_crosswalk)

        # ---------- 4. ВИЗУАЛИЗАЦИЯ ----------
        output_img = img.copy()

        if zebra_mask is not None:
            colored_mask = np.zeros_like(img)
            colored_mask[:, :, 1] = zebra_mask
            output_img = cv2.addWeighted(output_img, 0.7, colored_mask, 0.3, 0)
            if zebra_bbox:
                x, y, w, h = zebra_bbox
                cv2.rectangle(output_img, (x, y), (x + w, y + h), (0, 255, 0), 2)
                cv2.putText(output_img, 'ZEBRA CROSSING', (x, y - 10),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)
        else:
            cv2.putText(output_img, 'ZEBRA NOT FOUND', (10, 30),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 0, 255), 2)

        for p in pedestrians:
            x1, y1, x2, y2 = p['bbox']
            color = (0, 255, 0) if p in on_crosswalk else (255, 0, 0)
            cv2.rectangle(output_img, (x1, y1), (x2, y2), color, 2)
            cv2.putText(output_img, f'{p["conf"]:.2f}', (x1, y1-5),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.4, color, 1)

        # ---------- 5. СОХРАНЕНИЕ ----------
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        orig_name = f'orig_{ts}.jpg'
        proc_name = f'proc_{ts}.jpg'
        cv2.imwrite(os.path.join(STATIC_FOLDER, orig_name), img)
        cv2.imwrite(os.path.join(STATIC_FOLDER, proc_name), output_img)

        # ---------- 6. БАЗА ДАННЫХ ----------
        conn = sqlite3.connect('history.db')
        c = conn.cursor()
        c.execute("INSERT INTO detections (timestamp, original_file, processed_file, pedestrian_count, total_detected) VALUES (?, ?, ?, ?, ?)",
                  (datetime.now().isoformat(), orig_name, proc_name, count, total))
        conn.commit()
        conn.close()

        return jsonify({
            'count': count,
            'total_detected': total,
            'original_url': f'/static/{orig_name}',
            'image_url': f'/static/{proc_name}',
            'zebra_found': zebra_mask is not None
        })

    except Exception as e:
        print("Ошибка в /process:", traceback.format_exc())
        return jsonify({'error': str(e)}), 500

# ---------- СТАТИКА ----------
@app.route('/static/<path:filename>')
def static_files(filename):
    return send_from_directory(STATIC_FOLDER, filename)

# ---------- ОТЧЁТ PDF ----------
@app.route('/report')
def generate_report():
    try:
        conn = sqlite3.connect('history.db')
        c = conn.cursor()
        c.execute("SELECT id, timestamp, original_file, processed_file, pedestrian_count FROM detections ORDER BY timestamp DESC")
        rows = c.fetchall()
        conn.close()

        buffer = io.BytesIO()
        pdf = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4
        y = height - 20 * mm

        font_path = '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'
        pdfmetrics.registerFont(TTFont('DejaVu', font_path))

        # Заголовок (минималистичный)
        pdf.setFont('DejaVu', 12)
        pdf.drawString(20 * mm, y, "PEDESTRIAN MONITOR — LOG")
        y -= 6 * mm
        pdf.setFont('DejaVu', 8)
        pdf.drawString(20 * mm, y, f"generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        y -= 8 * mm

        # Линия-разделитель
        pdf.line(20 * mm, y, width - 20 * mm, y)
        y -= 8 * mm

        if not rows:
            pdf.setFont('DejaVu', 10)
            pdf.drawString(20 * mm, y, "No records")
        else:
            for row in rows:
                if y < 65 * mm:   # место для двух картинок + текст
                    pdf.showPage()
                    y = height - 20 * mm
                    pdf.setFont('DejaVu', 8)

                # Информация о записи
                pdf.setFont('DejaVu', 9)
                dt = datetime.fromisoformat(row[1])
                pdf.drawString(20 * mm, y, f"#{row[0]}  {dt.strftime('%d.%m.%Y %H:%M:%S')}  |  on zebra: {row[4]}")
                y -= 5 * mm

                orig_path = os.path.join(STATIC_FOLDER, row[2])
                proc_path = os.path.join(STATIC_FOLDER, row[3])

                if os.path.exists(orig_path) and os.path.exists(proc_path):
                    try:
                        img_orig = ImageReader(orig_path)
                        img_proc = ImageReader(proc_path)
                        img_w = 75 * mm
                        img_h = 55 * mm

                        # Оригинал слева
                        pdf.drawImage(img_orig, 20 * mm, y - img_h,
                                      width=img_w, height=img_h,
                                      preserveAspectRatio=True, anchor='c')
                        # Результат справа
                        pdf.drawImage(img_proc, 20 * mm + img_w + 8 * mm, y - img_h,
                                      width=img_w, height=img_h,
                                      preserveAspectRatio=True, anchor='c')

                        # Подписи под картинками (мелко, почти незаметно)
                        pdf.setFont('DejaVu', 6)
                        pdf.drawString(20 * mm + 5 * mm, y - img_h - 4 * mm, "before")
                        pdf.drawString(20 * mm + img_w + 8 * mm + 5 * mm, y - img_h - 4 * mm, "after")

                        y -= (img_h + 8 * mm)
                    except Exception as e:
                        pdf.drawString(20 * mm, y - 10 * mm, "[image error]")
                        y -= 12 * mm
                else:
                    pdf.drawString(20 * mm, y - 10 * mm, "[missing images]")
                    y -= 12 * mm

                # Разделитель между записями
                pdf.line(20 * mm, y, width - 20 * mm, y)
                y -= 6 * mm

        pdf.save()
        buffer.seek(0)
        return send_file(buffer, as_attachment=True, download_name='pedestrian_report.pdf', mimetype='application/pdf')

    except Exception as e:
        print("Ошибка в /report:", traceback.format_exc())
        return jsonify({'error': str(e)}), 500

# ---------- ОТЧЁТ EXCEL ----------
@app.route('/export_excel')
def export_excel():
    try:
        conn = sqlite3.connect('history.db')
        c = conn.cursor()
        c.execute("SELECT id, timestamp, original_file, processed_file, pedestrian_count FROM detections ORDER BY timestamp DESC")
        rows = c.fetchall()
        conn.close()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"

        # Заголовки (без жирности)
        headers = ["Date/Time", "On zebra", "Original", "Result"]
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = openpyxl.styles.Font(size=9, bold=False)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center')

        # Заполнение
        row_num = 2
        for row in rows:
            dt = datetime.fromisoformat(row[1])
            ws.cell(row=row_num, column=1, value=dt.strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row_num, column=2, value=row[4])

            # Вставляем изображения в колонки C и D
            orig_path = os.path.join(STATIC_FOLDER, row[2])
            proc_path = os.path.join(STATIC_FOLDER, row[3])

            if os.path.exists(orig_path):
                try:
                    img_orig = XLImage(orig_path)
                    img_orig.width = 100
                    img_orig.height = 75
                    ws.add_image(img_orig, f'C{row_num}')
                except:
                    ws.cell(row=row_num, column=3, value="err")
            else:
                ws.cell(row=row_num, column=3, value="—")

            if os.path.exists(proc_path):
                try:
                    img_proc = XLImage(proc_path)
                    img_proc.width = 100
                    img_proc.height = 75
                    ws.add_image(img_proc, f'D{row_num}')
                except:
                    ws.cell(row=row_num, column=4, value="err")
            else:
                ws.cell(row=row_num, column=4, value="—")

            row_num += 1

        # Настройка ширины и высоты
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 22   # под картинку
        ws.column_dimensions['D'].width = 22
        for r in range(2, row_num):
            ws.row_dimensions[r].height = 75   # под картинку

        # Убираем сетку (для чистоты)
        ws.sheet_view.showGridLines = False

        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return send_file(excel_file, as_attachment=True, download_name='pedestrian_history.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        print("Ошибка в /export_excel:", traceback.format_exc())
        return jsonify({'error': str(e)}), 500

# ---------- ЗАПУСК ----------
if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5555, debug=True)
